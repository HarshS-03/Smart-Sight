from django.utils import timezone
from datetime import timedelta, datetime
from django.http import StreamingHttpResponse, JsonResponse, HttpResponse
import time
import time as _time
import base64
import json
import os
import uuid
import smtplib
import threading
import requests
from email.message import EmailMessage
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.db.models.functions import TruncDate
from django.db.models import Count, Min, Max
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import cv2 as cv
import numpy as np
from ultralytics import YOLO
from dotenv import load_dotenv
from django.conf import settings
from .models import User, Person, PersonImage, RecognitionLog

load_dotenv()

# In-memory store for face login verifications (avoids StreamingHttpResponse session issues)
_face_login_verified = {}

# Cache for YOLO models
YOLO_MODELS = {}

def get_yolo_model(model_name):
    if model_name not in YOLO_MODELS:
        if model_name == 'yolov8n_onnx':
            model_path = 'app/models/nano/weights/best.onnx'
        elif model_name in ['yolov8n_pt', 'yolov8n']:
            model_path = 'app/models/nano/weights/best.pt'
        elif model_name == 'yolov8s_onnx':
            model_path = 'app/models/small/weights/best.onnx'
        elif model_name in ['yolov8s_pt', 'yolov8s']:
            model_path = 'app/models/small/weights/best.pt'
        else:
            model_path = f'app/models/{model_name}.pt'
            
        if not os.path.exists(model_path):
            # YOLO will automatically download if not found at exact path
            YOLO_MODELS[model_name] = YOLO(f'{model_name}.pt')
        else:
            if model_path.endswith('.onnx'):
                YOLO_MODELS[model_name] = YOLO(model_path, task='detect')
            else:
                YOLO_MODELS[model_name] = YOLO(model_path)
    return YOLO_MODELS[model_name]

# Create your views here.

def home(request):
    return render(request,'public/home.html')

def base(request):
    return render(request,'base.html')

def about(request):
    return render(request,'public/about.html')

@login_required
def detection(request):
    return render(request,'admin/detection.html')

@login_required
@user_passes_test(lambda u: u.is_staff)
def dataset(request):
    if request.method == 'POST':
        person_name = request.POST.get('name')
        images = request.FILES.getlist('images')
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            if person_name and images:
                person, created = Person.objects.get_or_create(name=person_name)
                for img in images:
                    PersonImage.objects.create(person=person, image=img)
                return JsonResponse({'status': 'success', 'message': f"Folder '{person_name}' updated."})
            elif person_name:
                Person.objects.get_or_create(name=person_name)
                return JsonResponse({'status': 'success', 'message': f"Folder '{person_name}' created."})
            return JsonResponse({'status': 'error', 'message': 'Missing data'}, status=400)

        if person_name and images:
            person, created = Person.objects.get_or_create(name=person_name)
            for img in images:
                PersonImage.objects.create(person=person, image=img)
            messages.success(request, f"Folder '{person_name}' updated with {len(images)} new images.")
        elif person_name:
            Person.objects.get_or_create(name=person_name)
            messages.success(request, f"Folder '{person_name}' created.")
        
        return redirect('dataset')

    persons = Person.objects.all().order_by('name')
    
    # Query all Unknown stranger logs containing physical images
    unknowns = RecognitionLog.objects.filter(status='UNKNOWN').exclude(image_path__isnull=True).exclude(image_path='').order_by('-timestamp')
    
    # Group unknown captures date-wise in local timezone
    unknowns_by_date = {}
    for log in unknowns:
        local_time = timezone.localtime(log.timestamp)
        date_str = local_time.strftime('%B %d, %Y')
        date_safe = local_time.strftime('%Y_%m_%d')
        if date_str not in unknowns_by_date:
            unknowns_by_date[date_str] = {
                'safe_id': date_safe,
                'logs': []
            }
        unknowns_by_date[date_str]['logs'].append({
            'id': log.id,
            'image_url': f"{settings.MEDIA_URL}{log.image_path}",
            'time_str': local_time.strftime('%I:%M:%S %p'),
            'confidence': round(log.confidence * 100, 1),
        })
        
    return render(request,'admin/dataset.html', {
        'persons': persons,
        'unknowns_by_date': unknowns_by_date
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def delete_person(request, person_id):
    person = get_object_or_404(Person, id=person_id)
    # Manually delete physical files for all images
    for img in person.images.all():
        if img.image and os.path.isfile(img.image.path):
            try:
                os.remove(img.image.path)
            except Exception:
                pass
    person.delete()
    messages.success(request, "Person and all associated images deleted successfully.")
    return redirect('dataset')

@login_required
@user_passes_test(lambda u: u.is_staff)
def delete_image(request, image_id):
    image = get_object_or_404(PersonImage, id=image_id)
    if image.image and os.path.isfile(image.image.path):
        try:
            os.remove(image.image.path)
        except Exception:
            pass
    image.delete()
    messages.success(request, "Training sample removed successfully.")
    return redirect('dataset')

@login_required
@user_passes_test(lambda u: u.is_staff)
def classify_unknowns(request):
    if request.method == 'POST':
        try:
            from app.face_classifier import classify_unknown_faces
            groups = classify_unknown_faces()
            return JsonResponse({'status': 'success', 'groups': groups})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Only POST method is allowed'}, status=405)

@login_required
@user_passes_test(lambda u: u.is_staff)
def assign_classified_group(request):
    if request.method == 'POST':
        try:
            from app.face_classifier import assign_group_to_person
            import json
            data = json.loads(request.body)
            group_images = data.get('group_images', [])
            person_name = data.get('person_name', '')
            
            if not group_images or not person_name:
                return JsonResponse({'status': 'error', 'message': 'Missing group images or person name.'}, status=400)
                
            success, msg = assign_group_to_person(group_images, person_name)
            if success:
                return JsonResponse({'status': 'success', 'message': msg})
            else:
                return JsonResponse({'status': 'error', 'message': msg}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Only POST method is allowed'}, status=405)

def login_user(request):
    if not request.session.session_key:
        request.session.create()
        
    if request.user.is_authenticated:
        return redirect('home')
        
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        if user is not None:
            login(request, user)
            messages.success(request, f"Welcome back, {u}!")
            return redirect('home')
        else:
            messages.error(request, "Invalid username or password.")
            return render(request, 'public/login.html')
    return render(request,'public/login.html')

def gen_face_login_frames(request):
    token = request.GET.get('token')
    print(f"[FaceLogin] Stream started for token: {token}")
    # Open the default server webcam
    cap = cv.VideoCapture(0)
    if cap is None or not cap.isOpened():
        print("[FaceLoginFeed] Error: Cannot open camera")
        # Yield an error frame
        error_frame = np.zeros((480, 640, 3), dtype=np.uint8)
        cv.putText(error_frame, "Camera Offline or In Use", (100, 240), 
                   cv.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2, cv.LINE_AA)
        ret, buffer = cv.imencode('.jpg', error_frame)
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
        return
        
    cap.set(cv.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv.CAP_PROP_FRAME_HEIGHT, 480)
    
    model = get_yolo_model('yolov8n')
    
    # Laser position for animated scanner line
    laser_y = 40
    laser_direction = 8
    
    success_frames_count = 0
    verified_user = None
    
    while True:
        success, frame = cap.read()
        if not success:
            break
            
        h, w, _ = frame.shape
        
        # Run YOLOv8 face recognition
        results = model(frame, conf=0.25, verbose=False)
        
        best_conf = 0.0
        recognized_name = None
        face_box = None
        
        for box in results[0].boxes:
            cls_id = int(box.cls[0])
            conf = float(box.conf[0])
            if conf > best_conf:
                best_conf = conf
                recognized_name = model.names[cls_id]
                face_box = [int(v) for v in box.xyxy[0]]
        
        # 1. Draw Biometric Corner HUD overlays directly on the video frame
        # Glowing brackets in blue (BGR: 253, 110, 13)
        hud_color = (253, 110, 13)
        # Top-left
        cv.line(frame, (40, 40), (70, 40), hud_color, 3)
        cv.line(frame, (40, 40), (40, 70), hud_color, 3)
        # Top-right
        cv.line(frame, (w-40, 40), (w-70, 40), hud_color, 3)
        cv.line(frame, (w-40, 40), (w-40, 70), hud_color, 3)
        # Bottom-left
        cv.line(frame, (40, h-40), (70, h-40), hud_color, 3)
        cv.line(frame, (40, h-40), (40, h-70), hud_color, 3)
        # Bottom-right
        cv.line(frame, (w-40, h-40), (w-70, h-40), hud_color, 3)
        cv.line(frame, (w-40, h-40), (w-40, h-70), hud_color, 3)
        
        # 2. Draw moving scan laser line
        cv.line(frame, (40, laser_y), (w-40, laser_y), (253, 110, 13), 2)
        # Add subtle glow to laser line
        cv.line(frame, (40, laser_y), (w-40, laser_y), (255, 180, 100), 1)
        laser_y += laser_direction
        if laser_y >= h - 40 or laser_y <= 40:
            laser_direction *= -1
            
        # 3. If face is detected, track it visually
        if face_box:
            x1, y1, x2, y2 = face_box
            
            # Match user in database
            is_admin = False
            try:
                user = User.objects.get(username__iexact=recognized_name)
                if user.is_staff or user.is_superuser:
                    is_admin = True
            except User.DoesNotExist:
                pass
                
            # Pick color: Green for admin >= 75%, Orange otherwise
            if best_conf >= 0.75 and is_admin:
                box_color = (84, 185, 25) # Green
                label_prefix = "[ADMIN]"
            elif is_admin:
                box_color = (0, 165, 255) # Orange (low confidence admin)
                label_prefix = "[LOW CONF]"
            else:
                box_color = (0, 0, 255) # Red (non-admin or unrecognized)
                label_prefix = "[BLOCKED]"
                
            # Draw rectangles
            cv.rectangle(frame, (x1, y1), (x2, y2), box_color, 2)
            
            # Draw label
            label = f"{label_prefix} {recognized_name} ({round(best_conf * 100, 1)}%)"
            cv.putText(frame, label, (x1, y1 - 10), cv.FONT_HERSHEY_SIMPLEX, 0.55, box_color, 2, cv.LINE_AA)
            
            # Verify conditions
            if best_conf >= 0.75 and is_admin:
                success_frames_count += 1
                if success_frames_count >= 2:
                    verified_user = user
            else:
                success_frames_count = max(0, success_frames_count - 1)
        else:
            success_frames_count = max(0, success_frames_count - 1)
            
        # If successfully verified, display ACCESS GRANTED overlay
        if verified_user:
            # Draw access granted visual HUD on frame
            cv.rectangle(frame, (0, h // 2 - 50), (w, h // 2 + 50), (10, 15, 30), -1) # Dark glassmorphic bar
            cv.rectangle(frame, (0, h // 2 - 50), (w, h // 2 + 50), (84, 185, 25), 2) # Green border
            cv.putText(frame, "BIOMETRIC IDENTITY VERIFIED", (w // 2 - 240, h // 2 - 10), 
                       cv.FONT_HERSHEY_SIMPLEX, 0.75, (255, 255, 255), 2, cv.LINE_AA)
            cv.putText(frame, f"ACCESS GRANTED - WELCOME {verified_user.username.upper()}", (w // 2 - 270, h // 2 + 25), 
                       cv.FONT_HERSHEY_SIMPLEX, 0.75, (84, 185, 25), 2, cv.LINE_AA)
            
            # Store verified user in memory dict (session won't work inside streaming generator)
            if token:
                _face_login_verified[token] = verified_user.username
                print(f"[FaceLogin] Verified username '{verified_user.username}' for token '{token}'")
            
            # Yield final confirmation frame and exit loop
            ret, buffer = cv.imencode('.jpg', frame)
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
            time.sleep(0.3) # Brief flash of success before frontend redirects
            break
            
        ret, buffer = cv.imencode('.jpg', frame)
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
        time.sleep(0.04) # cap ~25 FPS for lighter processing
        
    cap.release()

def face_login_feed(request):
    return StreamingHttpResponse(gen_face_login_frames(request), content_type='multipart/x-mixed-replace; boundary=frame')

@csrf_exempt
def face_login_check(request):
    token = request.GET.get('token')
    username = _face_login_verified.get(token) if token else None
    print(f"[FaceLoginCheck] Checking token: {token}, found username: {username}")
    if username:
        try:
            user = User.objects.get(username=username)
            login(request, user)
            messages.success(request, f"Welcome back, {user.username}! (Face ID Verified)")
            # Clear from memory store
            _face_login_verified.pop(token, None)
            return JsonResponse({'success': True, 'redirect_url': '/home/'})
        except User.DoesNotExist:
            _face_login_verified.pop(token, None)
    return JsonResponse({'success': False})

def logout_user(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('login')

def forgot_password(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        code = request.POST.get('code')
        new_password = request.POST.get('new_password')
        
        try:
            user = User.objects.get(username=username, code=code)
            if new_password:
                user.set_password(new_password)
                user.save()
                messages.success(request, "Access restored. Password has been updated.")
                return redirect('home')
            else:
                return render(request, 'public/forgot_password.html', {
                    'verified': True, 
                    'username': username, 
                    'code': code,
                    'success': 'Identity verified. Please set your new password.'
                })
        except User.DoesNotExist:
            messages.error(request, "Verification failed. Invalid username or security code.")
            
    return render(request, 'public/forgot_password.html')

# Global stats for live FPS and detection counts
_feed_stats = {"fps": 0, "persons": 0}

def _fix_camera_url(url):
    """Auto-fix common IP camera URL issues."""
    url = url.strip()
    # Replace https with http for local IP cameras
    if url.startswith('https://192.') or url.startswith('https://10.') or url.startswith('https://172.'):
        url = url.replace('https://', 'http://', 1)
        print(f"[VideoFeed] Auto-fixed: https -> http")
    
    # Auto-append /video when user gives just base URL (no path)
    if url.startswith('http://') or url.startswith('https://'):
        from urllib.parse import urlparse
        parsed = urlparse(url)
        if not parsed.path or parsed.path == '/':
            url = url.rstrip('/') + '/video'
            print(f"[VideoFeed] Auto-appended /video -> {url}")
    return url

def gen_frames(camera_src, model_name='yolov8n', orientation='normal'):
    global _feed_stats
    _feed_stats = {"fps": 0, "persons": 0}
    
    # Convert camera_src to int if it's a digit (for webcam index)
    if camera_src.isdigit():
        camera_src = int(camera_src)
    
    # Auto-fix URL for IP cameras
    if isinstance(camera_src, str):
        camera_src = _fix_camera_url(camera_src)
    
    print(f"[VideoFeed] Attempting to connect to camera: {camera_src}")
    
    cap = None
    # Use FFMPEG backend for network streams if it's a URL
    if isinstance(camera_src, str) and (camera_src.startswith('http') or camera_src.startswith('rtsp')):
        print(f"[VideoFeed] Detected network stream, trying FFMPEG backend...")
        for attempt in range(3):
            cap = cv.VideoCapture(camera_src, cv.CAP_FFMPEG)
            cap.set(cv.CAP_PROP_BUFFERSIZE, 3)
            if cap.isOpened():
                print(f"[VideoFeed] Connected on attempt {attempt + 1}")
                break
            cap.release()
            cap = None
            print(f"[VideoFeed] Attempt {attempt + 1} failed, retrying...")
            _time.sleep(1)
        
        # If FFMPEG failed, try default backend
        if cap is None or not cap.isOpened():
            print(f"[VideoFeed] FFMPEG failed, trying default backend...")
            cap = cv.VideoCapture(camera_src)
    else:
        cap = cv.VideoCapture(camera_src)
    
    if cap is None or not cap.isOpened():
        print(f"[VideoFeed] ERROR: Cannot connect to camera: {camera_src}")
        error_frame = np.zeros((480, 640, 3), dtype=np.uint8)
        cv.putText(error_frame, "Error: Cannot connect to Camera", (50, 220), 
                   cv.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2, cv.LINE_AA)
        cv.putText(error_frame, f"Src: {camera_src}", (50, 260), 
                   cv.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1, cv.LINE_AA)
        ret, buffer = cv.imencode('.jpg', error_frame)
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
        return

    # Try to set a lower resolution for better performance on CPU
    cap.set(cv.CAP_PROP_FRAME_WIDTH, 640)
    cap.set(cv.CAP_PROP_FRAME_HEIGHT, 480)

    model = get_yolo_model(model_name)
    
    # FPS tracking
    frame_count = 0
    fps_start_time = _time.time()

    continuous_detection_frames = 0

    while True:
        success, frame = cap.read()
        if not success:
            error_frame = np.zeros((480, 640, 3), dtype=np.uint8)
            cv.putText(error_frame, "Stream Lost", (200, 240), 
                       cv.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv.LINE_AA)
            ret, buffer = cv.imencode('.jpg', error_frame)
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
            break
        
        # Apply manual orientation rotation if specified
        if orientation == 'rot90_cw':
            frame = cv.rotate(frame, cv.ROTATE_90_CLOCKWISE)
        elif orientation == 'rot90_ccw':
            frame = cv.rotate(frame, cv.ROTATE_90_COUNTERCLOCKWISE)
        elif orientation == 'flip180':
            frame = cv.flip(frame, -1)
        
        # Run YOLOv8 detection
        results = model(frame, conf=0.5, verbose=False)
        
        # Determine if we are using the custom face model
        is_custom = (model_name in ['yolov8n_onnx', 'yolov8n_pt', 'yolov8n', 'yolov8s_onnx', 'yolov8s_pt', 'yolov8s'])
        
        # Count persons detected and track names
        person_count = 0
        person_detected = False
        max_confidence = 0.0
        detected_names_set = set()
        
        annotated_frame = frame.copy()
        
        for box in results[0].boxes:
            cls_id = int(box.cls[0])
            if is_custom or cls_id == 0:
                person_count += 1
                person_detected = True
                conf = float(box.conf[0])
                if conf > max_confidence:
                    max_confidence = conf
                
                # Identify name dynamically for this specific bounding box
                box_name = 'Unknown'
                if is_custom:
                    if conf >= 0.65:
                        box_name = model.names[cls_id]
                    detected_names_set.add(box_name)
                else:
                    detected_names_set.add('Unknown')
                
                # Draw Sci-Fi Bounding Box
                x1, y1, x2, y2 = [int(v) for v in box.xyxy[0]]
                
                if box_name != 'Unknown':
                    box_color = (84, 185, 25) # Green for known
                else:
                    box_color = (0, 0, 255) # Red for unknown/stranger
                
                # Translucent fill
                overlay = annotated_frame.copy()
                cv.rectangle(overlay, (x1, y1), (x2, y2), box_color, -1)
                cv.addWeighted(overlay, 0.15, annotated_frame, 0.85, 0, annotated_frame)
                
                # Glowing corner brackets
                length = min(30, int((x2 - x1) * 0.2)) # adjust bracket size
                thickness = max(2, int((x2 - x1) * 0.01))
                
                # Top-Left
                cv.line(annotated_frame, (x1, y1), (x1 + length, y1), box_color, thickness)
                cv.line(annotated_frame, (x1, y1), (x1, y1 + length), box_color, thickness)
                # Top-Right
                cv.line(annotated_frame, (x2, y1), (x2 - length, y1), box_color, thickness)
                cv.line(annotated_frame, (x2, y1), (x2, y1 + length), box_color, thickness)
                # Bottom-Left
                cv.line(annotated_frame, (x1, y2), (x1 + length, y2), box_color, thickness)
                cv.line(annotated_frame, (x1, y2), (x1, y2 - length), box_color, thickness)
                # Bottom-Right
                cv.line(annotated_frame, (x2, y2), (x2 - length, y2), box_color, thickness)
                cv.line(annotated_frame, (x2, y2), (x2, y2 - length), box_color, thickness)
                
                # Label
                label = f"{box_name} ({round(conf * 100, 1)}%)"
                (w, h), _ = cv.getTextSize(label, cv.FONT_HERSHEY_SIMPLEX, 0.5, 1)
                cv.rectangle(annotated_frame, (x1, y1 - 25), (x1 + w, y1), box_color, -1)
                cv.putText(annotated_frame, label, (x1, y1 - 8), cv.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1, cv.LINE_AA)
        
        # Update FPS
        frame_count += 1
        elapsed = _time.time() - fps_start_time
        if elapsed >= 1.0:
            _feed_stats["fps"] = round(frame_count / elapsed, 1)
            _feed_stats["persons"] = person_count
            frame_count = 0
            fps_start_time = _time.time()
        else:
            _feed_stats["persons"] = person_count
                
        ret, buffer = cv.imencode('.jpg', annotated_frame)
        if not ret:
            continue
            
        frame_bytes = buffer.tobytes()
        
        # Trigger alert if person detected continuously
        if person_detected:
            continuous_detection_frames += 1
        else:
            continuous_detection_frames = max(0, continuous_detection_frames - 2)
            
        if continuous_detection_frames == 35:
            # Consolidate all unique detected names in this frame
            if detected_names_set:
                detected_name = ", ".join(sorted(list(detected_names_set)))
                # If "Unknown" is in the names, overall status is UNKNOWN (Intruder Warning Priority!)
                is_known = ("Unknown" not in detected_names_set)
            else:
                detected_name = 'Unknown'
                is_known = False
                
            # Encode clean unannotated frame for self-learning dataset gathering
            ret_clean, clean_buffer = cv.imencode('.jpg', frame)
            clean_frame_bytes = clean_buffer.tobytes() if ret_clean else None
            
            threading.Thread(
                target=send_alerts, 
                args=(frame_bytes, detected_name, is_known, person_count, max_confidence),
                kwargs={"clean_frame_bytes": clean_frame_bytes}
            ).start()
            
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
    
    cap.release()
    _feed_stats = {"fps": 0, "persons": 0}

# --- Telegram OK/Cancel Approval + 1-Hour Cooldown Flow ---

_pending_alerts = {}
_unknown_cooldowns = {}
UNKNOWN_COOLDOWN_SECONDS = 3600  # 1 hour
PENDING_ALERT_TIMEOUT = 600  # 10 minutes
_polling_started = False

def handle_admin_response(action, alert_id, cb_id, chat_id, message_id):
    telegram_bot_api = os.environ.get("telegram_bot_api")
    if not telegram_bot_api:
        return
        
    alert = _pending_alerts.pop(alert_id, None)
    if not alert:
        try:
            requests.post(
                f"https://api.telegram.org/bot{telegram_bot_api}/answerCallbackQuery",
                json={"callback_query_id": cb_id, "text": "This alert has already expired or been handled."}
            )
        except Exception as e:
            print(f"[Telegram Polling] Error answering callback query: {e}")
        return

    person_name = alert["person_name"]
    confidence = alert["confidence"]
    frame_bytes = alert["frame_bytes"]
    timestamp = alert["timestamp"]
    
    if action == "ok":
        relative_image_path = None
        if frame_bytes:
            try:
                unknown_dir = os.path.join(settings.MEDIA_ROOT, 'unknown')
                os.makedirs(unknown_dir, exist_ok=True)
                filename = f"unknown_{timestamp.strftime('%Y%m%d_%H%M%S')}_{alert_id}.jpg"
                file_path = os.path.join(unknown_dir, filename)
                with open(file_path, 'wb') as f:
                    f.write(frame_bytes)
                relative_image_path = f"unknown/{filename}"
                print(f"[Telegram Polling] Admin approved: saved unknown frame locally to: {relative_image_path}")
            except Exception as e:
                print(f"[Telegram Polling] Error saving unknown frame locally: {e}")
                
        try:
            RecognitionLog.objects.create(
                person_name=person_name,
                confidence=confidence,
                status='UNKNOWN',
                image_path=relative_image_path,
            )
            # Activate cooldown
            _unknown_cooldowns[person_name] = time.time()
            print(f"[Telegram Polling] Approved and logged: {person_name}, cooldown activated.")
        except Exception as e:
            print(f"[Telegram Polling] DB save error: {e}")
            
        new_caption = alert["details"] + "\n\n\u2705 <b>Status: APPROVED & LOGGED</b>"
        try:
            requests.post(
                f"https://api.telegram.org/bot{telegram_bot_api}/editMessageCaption",
                json={
                    "chat_id": chat_id,
                    "message_id": message_id,
                    "caption": new_caption,
                    "parse_mode": "HTML"
                }
            )
            requests.post(
                f"https://api.telegram.org/bot{telegram_bot_api}/answerCallbackQuery",
                json={"callback_query_id": cb_id, "text": "Alert Approved & Logged!"}
            )
        except Exception as e:
            print(f"[Telegram Polling] Error editing Telegram caption: {e}")
            
    elif action == "cancel":
        print(f"[Telegram Polling] Admin canceled alert for {person_name}. Discarding.")
        new_caption = alert["details"] + "\n\n\u274c <b>Status: CANCELED & DISCARDED</b>"
        try:
            requests.post(
                f"https://api.telegram.org/bot{telegram_bot_api}/editMessageCaption",
                json={
                    "chat_id": chat_id,
                    "message_id": message_id,
                    "caption": new_caption,
                    "parse_mode": "HTML"
                }
            )
            requests.post(
                f"https://api.telegram.org/bot{telegram_bot_api}/answerCallbackQuery",
                json={"callback_query_id": cb_id, "text": "Alert Canceled & Discarded."}
            )
        except Exception as e:
            print(f"[Telegram Polling] Error editing Telegram caption: {e}")

def check_alert_timeouts():
    telegram_bot_api = os.environ.get("telegram_bot_api")
    if not telegram_bot_api:
        return
        
    now = time.time()
    expired_ids = [aid for aid, alert in _pending_alerts.items() if now - alert["saved_at"] > PENDING_ALERT_TIMEOUT]
    
    for aid in expired_ids:
        alert = _pending_alerts.pop(aid, None)
        if not alert:
            continue
            
        person_name = alert["person_name"]
        confidence = alert["confidence"]
        frame_bytes = alert["frame_bytes"]
        timestamp = alert["timestamp"]
        chat_id = alert.get("chat_id")
        message_id = alert.get("message_id")
        
        relative_image_path = None
        if frame_bytes:
            try:
                unknown_dir = os.path.join(settings.MEDIA_ROOT, 'unknown')
                os.makedirs(unknown_dir, exist_ok=True)
                filename = f"unknown_{timestamp.strftime('%Y%m%d_%H%M%S')}_{aid}.jpg"
                file_path = os.path.join(unknown_dir, filename)
                with open(file_path, 'wb') as f:
                    f.write(frame_bytes)
                relative_image_path = f"unknown/{filename}"
            except Exception as e:
                print(f"[Timeout Engine] Error saving frame locally: {e}")
                
        try:
            RecognitionLog.objects.create(
                person_name=person_name,
                confidence=confidence,
                status='UNKNOWN',
                image_path=relative_image_path,
            )
            print(f"[Timeout Engine] Auto-saved alert due to 10-minute timeout: {person_name}")
        except Exception as e:
            print(f"[Timeout Engine] DB save error: {e}")
            
        if chat_id and message_id:
            new_caption = alert["details"] + "\n\n\u23f1\ufe0f <b>Status: AUTO-SAVED (TIMEOUT)</b>"
            try:
                requests.post(
                    f"https://api.telegram.org/bot{telegram_bot_api}/editMessageCaption",
                    json={
                        "chat_id": chat_id,
                        "message_id": message_id,
                        "caption": new_caption,
                        "parse_mode": "HTML"
                    }
                )
            except Exception as e:
                print(f"[Timeout Engine] Error editing message caption: {e}")

def _start_telegram_polling():
    global _polling_started
    if _polling_started:
        return
    _polling_started = True
    
    def poll_loop():
        telegram_bot_api = os.environ.get("telegram_bot_api")
        if not telegram_bot_api:
            print("[Telegram Listener] No telegram_bot_api token found. Listener not starting.")
            return
        
        try:
            requests.get(f"https://api.telegram.org/bot{telegram_bot_api}/deleteWebhook")
        except Exception as e:
            print(f"[Telegram Listener] Error deleting webhook: {e}")

        offset = 0
        print("[Telegram Listener] Started polling for callback queries...")
        
        while True:
            try:
                url = f"https://api.telegram.org/bot{telegram_bot_api}/getUpdates"
                params = {"timeout": 10, "offset": offset}
                response = requests.get(url, params=params, timeout=15)
                if response.status_code == 200:
                    data = response.json()
                    if data.get("ok"):
                        for update in data.get("result", []):
                            update_id = update["update_id"]
                            offset = update_id + 1
                            
                            if "callback_query" in update:
                                cb = update["callback_query"]
                                cb_id = cb["id"]
                                cb_data = cb.get("data", "")
                                message = cb.get("message", {})
                                chat_id = message.get("chat", {}).get("id")
                                message_id = message.get("message_id")
                                
                                if cb_data.startswith("ok_") or cb_data.startswith("cancel_"):
                                    action, alert_id = cb_data.split("_", 1)
                                    handle_admin_response(action, alert_id, cb_id, chat_id, message_id)
                
                check_alert_timeouts()
                
            except Exception as e:
                print(f"[Telegram Listener] Polling error: {e}")
            time.sleep(2)
            
    threading.Thread(target=poll_loop, daemon=True).start()

# Start polling listener automatically when module is loaded
_start_telegram_polling()

def send_alerts(*args, **kwargs):
    # Default values
    frame_bytes = None
    person_name = 'Unknown'
    is_known = False
    person_count = 1
    confidence = 0.0

    if len(args) == 5:
        frame_bytes, person_name, is_known, person_count, confidence = args
    elif len(args) == 3:
        frame_bytes, person_count, confidence = args
    elif len(args) >= 1:
        frame_bytes = args[0]
        if len(args) > 1:
            person_count = args[1]
        if len(args) > 2:
            confidence = args[2]

    # Map kwargs if any
    if 'frame_bytes' in kwargs: frame_bytes = kwargs['frame_bytes']
    if 'person_name' in kwargs: person_name = kwargs['person_name']
    if 'is_known' in kwargs: is_known = kwargs['is_known']
    if 'person_count' in kwargs: person_count = kwargs['person_count']
    if 'confidence' in kwargs: confidence = kwargs['confidence']

    timestamp = datetime.now()
    timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')
    confidence_pct = round(confidence * 100, 1)
    status_str = 'KNOWN' if is_known else 'UNKNOWN'

    telegram_bot_api = os.environ.get("telegram_bot_api")
    telegram_chat_id = os.environ.get("telegram_chat_id")

    if not telegram_bot_api or not telegram_chat_id:
        print(f"[Alert System Warning] Telegram credentials missing! bot_api={telegram_bot_api}, chat_id={telegram_chat_id}")

    if not is_known:
        # Check 1-hour cooldown
        current_time = time.time()
        cooldown_active = False
        if person_name in _unknown_cooldowns:
            time_passed = current_time - _unknown_cooldowns[person_name]
            if time_passed < UNKNOWN_COOLDOWN_SECONDS:
                cooldown_active = True
            else:
                del _unknown_cooldowns[person_name]

        if cooldown_active:
            # Path A: Cooldown Active -> Auto-save image to disk + DB, brief Telegram notification
            relative_image_path = None
            if frame_bytes:
                try:
                    unknown_dir = os.path.join(settings.MEDIA_ROOT, 'unknown')
                    os.makedirs(unknown_dir, exist_ok=True)
                    filename = f"unknown_{timestamp.strftime('%Y%m%d_%H%M%S')}_{int(time.time())}.jpg"
                    file_path = os.path.join(unknown_dir, filename)
                    with open(file_path, 'wb') as f:
                        f.write(frame_bytes)
                    relative_image_path = f"unknown/{filename}"
                except Exception as e:
                    print(f"[Alert Cooldown] Error saving unknown frame: {e}")

            try:
                RecognitionLog.objects.create(
                    person_name=person_name,
                    confidence=confidence,
                    status='UNKNOWN',
                    image_path=relative_image_path,
                )
                print(f"[Alert Cooldown] Cooldown active. Auto-saved: {person_name}")
            except Exception as e:
                print(f"[Alert Cooldown] DB save error: {e}")

            if telegram_bot_api and telegram_chat_id:
                try:
                    text_msg = (
                        f"\u23f1\ufe0f <b>Same Person Detected (Cooldown Active)</b>\n"
                        f"\U0001f464 Person: {person_name}\n"
                        f"\U0001f552 Time: {timestamp_str}\n"
                        f"\u2705 <b>Auto-saved in DB & Unknown Folder</b>"
                    )
                    url = f"https://api.telegram.org/bot{telegram_bot_api}/sendMessage"
                    requests.post(url, json={
                        "chat_id": telegram_chat_id,
                        "text": text_msg,
                        "parse_mode": "HTML"
                    })
                except Exception as e:
                    print(f"[Alert Cooldown] Telegram text error: {e}")
        else:
            # Path B: No Cooldown -> Save image locally in memory, send Telegram photo with OK/Cancel
            alert_id = uuid.uuid4().hex[:12]
            
            status_icon = "\U0001f6a8"
            alert_details = (
                f"{status_icon} <b>Security Alert - Smart Sight</b>\n"
                f"\n"
                f"\U0001f464 <b>Person:</b> {person_name}\n"
                f"\U0001f465 <b>Count:</b> {person_count} person(s) detected\n"
                f"\U0001f3af <b>Confidence:</b> {confidence_pct}%\n"
                f"\U0001f552 <b>Time:</b> {timestamp_str}\n"
                f"\U0001f4cc <b>Status:</b> {status_str}"
            )

            # Store in pending alerts (do NOT save to database yet)
            _pending_alerts[alert_id] = {
                "frame_bytes": frame_bytes,
                "person_name": person_name,
                "confidence": confidence,
                "timestamp": timestamp,
                "details": alert_details,
                "saved_at": current_time,
            }

            if telegram_bot_api and telegram_chat_id:
                try:
                    url = f"https://api.telegram.org/bot{telegram_bot_api}/sendPhoto"
                    files = {'photo': ('alert.jpg', frame_bytes, 'image/jpeg')}
                    reply_markup = {
                        "inline_keyboard": [
                            [
                                {"text": "OK \u2705", "callback_data": f"ok_{alert_id}"},
                                {"text": "Cancel \u274c", "callback_data": f"cancel_{alert_id}"}
                            ]
                        ]
                    }
                    data = {
                        'chat_id': telegram_chat_id,
                        'caption': alert_details,
                        'parse_mode': 'HTML',
                        'reply_markup': json.dumps(reply_markup)
                    }
                    response = requests.post(url, files=files, data=data)
                    if response.status_code == 200:
                        resp_json = response.json()
                        if resp_json.get("ok"):
                            msg_id = resp_json["result"]["message_id"]
                            if alert_id in _pending_alerts:
                                _pending_alerts[alert_id]["message_id"] = msg_id
                                _pending_alerts[alert_id]["chat_id"] = telegram_chat_id
                        print("[Alert] Telegram OK/Cancel alert sent successfully.")
                    else:
                        print(f"[Alert] Telegram API error: {response.text}")
                except Exception as e:
                    print(f"[Alert] Telegram send error: {e}")
    else:
        # Known Person: save immediately to database (no cooldown needed, no Telegram alert sent for known)
        try:
            RecognitionLog.objects.create(
                person_name=person_name,
                confidence=confidence,
                status=status_str,
                image_path=None,
            )
            print(f"[Alert] Saved Known to DB: {person_name}, confidence={confidence_pct}%")
            
            # Auto-save clean frame into the Person's dataset for more training data (Self-Learning)
            clean_frame_bytes = kwargs.get('clean_frame_bytes')
            if clean_frame_bytes:
                names = [n.strip() for n in person_name.split(',')]
                for name in names:
                    if name and name.lower() != 'unknown':
                        person = Person.objects.filter(name__iexact=name).first()
                        if person:
                            dataset_dir = os.path.join(settings.MEDIA_ROOT, 'dataset', person.name)
                            os.makedirs(dataset_dir, exist_ok=True)
                            
                            # Create a unique file name
                            filename = f"auto_{timestamp.strftime('%Y%m%d_%H%M%S')}_{int(time.time())}.jpg"
                            file_path = os.path.join(dataset_dir, filename)
                            
                            with open(file_path, 'wb') as f:
                                f.write(clean_frame_bytes)
                                
                            relative_path = f"dataset/{person.name}/{filename}"
                            PersonImage.objects.create(person=person, image=relative_path)
                            print(f"[Self-Learning Engine] Automatically saved training sample for {person.name}: {relative_path}")
        except Exception as e:
            print(f"[Self-Learning Engine] Error saving auto-training image: {e}")

def video_feed(request):
    # Get camera source from query params, default to 0 (webcam)
    src = request.GET.get('src', '0')
    model_name = request.GET.get('model', 'yolov8n')
    orientation = request.GET.get('orientation', 'normal')
    return StreamingHttpResponse(gen_frames(src, model_name, orientation), content_type='multipart/x-mixed-replace; boundary=frame')

def video_stats(request):
    return JsonResponse({"fps": _feed_stats["fps"], "faces": _feed_stats["persons"]})

@login_required
@user_passes_test(lambda u: u.is_staff)
def reports_view(request):
    search_query = request.GET.get('search', '').strip()
    status_query = request.GET.get('status', '')
    timeframe_query = request.GET.get('timeframe', 'all')
    
    # 1. Base log querying for the daily logs table
    logs = RecognitionLog.objects.all().order_by('-timestamp')
    if search_query:
        logs = logs.filter(person_name__icontains=search_query)
    if status_query:
        logs = logs.filter(status=status_query)
        
    if timeframe_query == 'today':
        logs = logs.filter(timestamp__date=timezone.now().date())
    elif timeframe_query == 'week':
        logs = logs.filter(timestamp__date__gte=timezone.now().date() - timedelta(days=7))
    elif timeframe_query == 'month':
        logs = logs.filter(timestamp__date__gte=timezone.now().date() - timedelta(days=30))
        
    # 2. Derive Daily Entry/Exit Reports (grouped by date, person, and status)
    daily_reports = logs.annotate(date=TruncDate('timestamp')).values('date', 'person_name', 'status').annotate(
        entry_time=Min('timestamp'),
        exit_time=Max('timestamp'),
        frequency=Count('id'),
        max_confidence=Max('confidence')
    ).order_by('-date', '-entry_time')
    
    # Pagination for daily aggregated reports
    paginator = Paginator(daily_reports, 10)  # 10 daily reports per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 3. Automatic Frequent Persons Analytics (Weekly & Monthly Frequency)
    end_date = timezone.now().date()
    start_month = end_date - timedelta(days=29)
    start_week = end_date - timedelta(days=6)
    
    # Query top 5 most frequent persons (excluding None for unknown names, or grouping them cleanly)
    frequent_persons_query = RecognitionLog.objects.values('person_name', 'status').annotate(
        total_count=Count('id')
    ).order_by('-total_count')[:5]
    
    frequent_persons = []
    chart_datasets = []
    
    colors = [
        {'border': '#0d6efd', 'bg': 'rgba(13, 110, 253, 0.1)'}, # Blue
        {'border': '#198754', 'bg': 'rgba(25, 135, 84, 0.1)'},  # Green
        {'border': '#dc3545', 'bg': 'rgba(220, 53, 69, 0.1)'},  # Red
        {'border': '#ffc107', 'bg': 'rgba(255, 193, 7, 0.1)'},  # Yellow
        {'border': '#0dcaf0', 'bg': 'rgba(13, 202, 240, 0.1)'}  # Cyan
    ]
    
    last_7_days = [end_date - timedelta(days=i) for i in range(6, -1, -1)]
    chart_labels = [d.strftime('%a (%d/%m)') for d in last_7_days]
    
    for idx, item in enumerate(frequent_persons_query):
        name = item['person_name']
        status = item['status']
        name_display = name if name else "Unknown Person"
        
        # Calculate weekly frequency (last 7 days)
        weekly_count = RecognitionLog.objects.filter(
            person_name=name,
            status=status,
            timestamp__date__range=[start_week, end_date]
        ).count()
        
        # Calculate monthly frequency (last 30 days)
        monthly_count = RecognitionLog.objects.filter(
            person_name=name,
            status=status,
            timestamp__date__range=[start_month, end_date]
        ).count()
        
        # Calculate daily counts for sparkline/chart
        daily_counts = RecognitionLog.objects.filter(
            person_name=name,
            status=status,
            timestamp__date__range=[start_week, end_date]
        ).annotate(date=TruncDate('timestamp')).values('date').annotate(count=Count('id'))
        
        daily_map = {d: 0 for d in last_7_days}
        for dc in daily_counts:
            dc_date = dc['date']
            if isinstance(dc_date, str):
                dc_date = datetime.strptime(dc_date, '%Y-%m-%d').date()
            if dc_date in daily_map:
                daily_map[dc_date] = dc['count']
                
        sparkline_values = [daily_map[d] for d in last_7_days]
        
        # Last seen record
        last_seen_log = RecognitionLog.objects.filter(
            person_name=name,
            status=status
        ).order_by('-timestamp').first()
        
        last_seen_time = last_seen_log.timestamp if last_seen_log else None
        
        frequent_persons.append({
            'name': name_display,
            'status': status,
            'total_count': item['total_count'],
            'weekly_count': weekly_count,
            'monthly_count': monthly_count,
            'last_seen': last_seen_time,
            'sparkline_values': sparkline_values
        })
        
        # Add to comparison chart datasets
        color = colors[idx % len(colors)]
        chart_datasets.append({
            'label': name_display,
            'data': sparkline_values,
            'borderColor': color['border'],
            'backgroundColor': color['bg'],
            'borderWidth': 3,
            'pointBackgroundColor': color['border'],
            'pointRadius': 4,
            'tension': 0.3,
            'fill': False
        })
        
    # Stats helpers
    total_detections = RecognitionLog.objects.count()
    known_detections = RecognitionLog.objects.filter(status='KNOWN').count()
    unknown_detections = RecognitionLog.objects.filter(status='UNKNOWN').count()
    
    # Get all unique person names from logs for advanced download filtering
    unique_names = RecognitionLog.objects.exclude(person_name__in=[None, '', 'Unknown']).values_list('person_name', flat=True).distinct().order_by('person_name')
    unique_names_list = list(unique_names)
    
    import json
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_query': status_query,
        'timeframe_query': timeframe_query,
        'total_detections': total_detections,
        'known_detections': known_detections,
        'unknown_detections': unknown_detections,
        'frequent_persons': frequent_persons,
        'unique_names_list': unique_names_list,
        'chart_labels': json.dumps(chart_labels),
        'chart_datasets': json.dumps(chart_datasets),
    }
    return render(request, 'admin/reports.html', context)

def _fill_excel_worksheet(ws, reports, title_text):
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Add title row
    ws.append([title_text])
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid") # Django Blue
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Add empty spacing row
    ws.append([])
    ws.row_dimensions[2].height = 15
    
    # Add Headers
    headers = ["Date", "Person Name", "Classification", "Entry Time (First Seen)", "Exit Time (Last Seen)", "Frequency (Detections)", "Max Confidence"]
    ws.append(headers)
    ws.row_dimensions[3].height = 25
    
    # Style Headers
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark slate
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for col_num in range(1, 8):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
    # Fill Data
    row_idx = 4
    for rep in reports:
        # Format Date
        rep_date = rep['date']
        if isinstance(rep_date, str):
            date_str = rep_date
        else:
            date_str = rep_date.strftime('%Y-%m-%d')
            
        name = rep['person_name'] if rep['person_name'] else "Unknown Person"
        status_disp = "Known" if rep['status'] == "KNOWN" else "Unknown"
        entry_local = timezone.localtime(rep['entry_time'])
        exit_local  = timezone.localtime(rep['exit_time'])
        entry_str = entry_local.strftime('%H:%M:%S')
        exit_str  = exit_local.strftime('%H:%M:%S')
        freq = rep['frequency']
        max_conf = f"{rep['max_confidence'] * 100:.1f}%" if rep['max_confidence'] <= 1.0 else f"{rep['max_confidence']:.1f}%"
        
        ws.append([date_str, name, status_disp, entry_str, exit_str, freq, max_conf])
        ws.row_dimensions[row_idx].height = 20
        
        # Style cells in this row
        row_fill = PatternFill(start_color="F8FAFC" if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")
        status_color = "15803D" if rep['status'] == "KNOWN" else "B91C1C" # Green vs Red
        status_font = Font(name="Segoe UI", size=10, bold=True, color=status_color)
        
        for col_num in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.fill = row_fill
            cell.border = thin_border
            cell.font = Font(name="Segoe UI", size=10) if col_num != 3 else status_font
            
            # Alignments
            if col_num in [1, 3, 4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        row_idx += 1
        
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
                break
                
        if not col_letter:
            continue
            
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_reports_excel(request):
    search_query = request.GET.get('search', '').strip()
    status_query = request.GET.get('status', '')
    time_range = request.GET.get('time_range', 'all')
    
    # Advanced Filter Queries
    export_date = request.GET.get('export_date')
    export_time_start = request.GET.get('export_time_start')
    export_time_end = request.GET.get('export_time_end')
    export_person = request.GET.get('export_person', '').strip()
    export_mode = request.GET.get('export_mode', 'combined') # 'combined' or 'separated'
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    logs = RecognitionLog.objects.all().order_by('-timestamp')
    if search_query:
        logs = logs.filter(person_name__icontains=search_query)
    if status_query:
        logs = logs.filter(status=status_query)
        
    now_local = timezone.localtime(timezone.now())
    today_local = now_local.date()
    
    title_suffix = "Report"
    filename = "Surveillance_Report.xlsx"
    
    # Standard time filter dropdown ranges
    if time_range == 'daily':
        logs = logs.filter(timestamp__date=today_local)
        title_suffix = "Daily Report"
        filename = "Surveillance_Daily_Report.xlsx"
    elif time_range == 'weekly':
        logs = logs.filter(timestamp__date__gte=today_local - timedelta(days=7))
        title_suffix = "Weekly Report"
        filename = "Surveillance_Weekly_Report.xlsx"
    elif time_range == 'monthly':
        logs = logs.filter(timestamp__date__gte=today_local - timedelta(days=30))
        title_suffix = "Monthly Report"
        filename = "Surveillance_Monthly_Report.xlsx"
    elif time_range == 'yearly':
        logs = logs.filter(timestamp__date__gte=today_local - timedelta(days=365))
        title_suffix = "Yearly Report"
        filename = "Surveillance_Yearly_Report.xlsx"
    elif time_range == 'all':
        title_suffix = "All-Time Report"
        filename = "Surveillance_All_Time_Report.xlsx"
    elif time_range == 'custom' and start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__gte=start, timestamp__date__lte=end)
            title_suffix = f"Custom Report ({start_date} to {end_date})"
            filename = f"Surveillance_Custom_Report_{start_date}_to_{end_date}.xlsx"
        except ValueError:
            title_suffix = "Custom Report"
            filename = "Surveillance_Custom_Report.xlsx"
            
    # Apply Advanced Single Filters
    if export_date:
        try:
            target_date = datetime.strptime(export_date, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date=target_date)
            title_suffix += f" - {export_date}"
            filename = f"Surveillance_Report_{export_date}.xlsx"
        except ValueError:
            pass
            
    if export_time_start and export_time_end:
        try:
            start_t = datetime.strptime(export_time_start, '%H:%M').time()
            end_t = datetime.strptime(export_time_end, '%H:%M').time()
            logs = logs.filter(timestamp__time__range=(start_t, end_t))
            title_suffix += f" ({export_time_start}-{export_time_end})"
        except ValueError:
            pass
            
    if export_person and export_person != 'All':
        if export_person.lower() == 'unknown':
            logs = logs.filter(status='UNKNOWN')
            title_suffix += " - Unknown Persons"
        else:
            logs = logs.filter(person_name__iexact=export_person)
            title_suffix += f" - {export_person}"
        filename = f"Surveillance_Report_{export_person}.xlsx"
        
    daily_reports = logs.annotate(date=TruncDate('timestamp')).values('date', 'person_name', 'status').annotate(
        entry_time=Min('timestamp'),
        exit_time=Max('timestamp'),
        frequency=Count('id'),
        max_confidence=Max('confidence')
    ).order_by('-date', '-entry_time')
    
    wb = Workbook()
    
    if export_mode == 'separated':
        # Group by person name
        unique_persons = set()
        for rep in daily_reports:
            p_name = rep['person_name'] if rep['person_name'] else "Unknown Person"
            unique_persons.add(p_name)
            
        if not unique_persons:
            ws = wb.active
            ws.title = "No Records"
            ws.append(["No records found matching filters."])
        else:
            for i, p_name in enumerate(sorted(list(unique_persons))):
                sheet_title = p_name[:30].replace('[','').replace(']','').replace('*','').replace('?','').replace(':','').replace('/','').replace('\\','')
                if i == 0:
                    ws = wb.active
                    ws.title = sheet_title
    ws.row_dimensions[3].height = 25
    
    # Style Headers
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark slate
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for col_num in range(1, 8):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
    # Fill Data
    row_idx = 4
    for rep in reports:
        # Format Date
        rep_date = rep['date']
        if isinstance(rep_date, str):
            date_str = rep_date
        else:
            date_str = rep_date.strftime('%Y-%m-%d')
            
        name = rep['person_name'] if rep['person_name'] else "Unknown Person"
        status_disp = "Known" if rep['status'] == "KNOWN" else "Unknown"
        entry_local = timezone.localtime(rep['entry_time'])
        exit_local  = timezone.localtime(rep['exit_time'])
        entry_str = entry_local.strftime('%H:%M:%S')
        exit_str  = exit_local.strftime('%H:%M:%S')
        freq = rep['frequency']
        max_conf = f"{rep['max_confidence'] * 100:.1f}%" if rep['max_confidence'] <= 1.0 else f"{rep['max_confidence']:.1f}%"
        
        ws.append([date_str, name, status_disp, entry_str, exit_str, freq, max_conf])
        ws.row_dimensions[row_idx].height = 20
        
        # Style cells in this row
        row_fill = PatternFill(start_color="F8FAFC" if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")
        status_color = "15803D" if rep['status'] == "KNOWN" else "B91C1C" # Green vs Red
        status_font = Font(name="Segoe UI", size=10, bold=True, color=status_color)
        
        for col_num in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.fill = row_fill
            cell.border = thin_border
            cell.font = Font(name="Segoe UI", size=10) if col_num != 3 else status_font
            
            # Alignments
            if col_num in [1, 3, 4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        row_idx += 1
        
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
                break
                
        if not col_letter:
            continue
            
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

@login_required
@user_passes_test(lambda u: u.is_staff)
def export_reports_excel(request):
    search_query = request.GET.get('search', '').strip()
    status_query = request.GET.get('status', '')
    time_range = request.GET.get('time_range', 'all')
    
    # Advanced Filter Queries
    export_date = request.GET.get('export_date')
    export_time_start = request.GET.get('export_time_start')
    export_time_end = request.GET.get('export_time_end')
    export_person = request.GET.get('export_person', '').strip()
    export_mode = request.GET.get('export_mode', 'combined') # 'combined' or 'separated'
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    logs = RecognitionLog.objects.all().order_by('-timestamp')
    if search_query:
        logs = logs.filter(person_name__icontains=search_query)
    if status_query:
        logs = logs.filter(status=status_query)
        
    now_local = timezone.localtime(timezone.now())
    today_local = now_local.date()
    
    title_suffix = "Report"
    filename = "Surveillance_Report.xlsx"
    
    # Standard time filter dropdown ranges
    if time_range == 'daily':
        logs = logs.filter(timestamp__date=today_local)
        title_suffix = "Daily Report"
        filename = "Surveillance_Daily_Report.xlsx"
    elif time_range == 'weekly':
        logs = logs.filter(timestamp__date__gte=today_local - timedelta(days=7))
        title_suffix = "Weekly Report"
        filename = "Surveillance_Weekly_Report.xlsx"
    elif time_range == 'monthly':
        logs = logs.filter(timestamp__date__gte=today_local - timedelta(days=30))
        title_suffix = "Monthly Report"
        filename = "Surveillance_Monthly_Report.xlsx"
    elif time_range == 'yearly':
        logs = logs.filter(timestamp__date__gte=today_local - timedelta(days=365))
        title_suffix = "Yearly Report"
        filename = "Surveillance_Yearly_Report.xlsx"
    elif time_range == 'all':
        title_suffix = "All-Time Report"
        filename = "Surveillance_All_Time_Report.xlsx"
    elif time_range == 'custom' and start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__gte=start, timestamp__date__lte=end)
            title_suffix = f"Custom Report ({start_date} to {end_date})"
            filename = f"Surveillance_Custom_Report_{start_date}_to_{end_date}.xlsx"
        except ValueError:
            title_suffix = "Custom Report"
            filename = "Surveillance_Custom_Report.xlsx"
            
    # Apply Advanced Single Filters
    if export_date:
        try:
            target_date = datetime.strptime(export_date, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date=target_date)
            title_suffix += f" - {export_date}"
            filename = f"Surveillance_Report_{export_date}.xlsx"
        except ValueError:
            pass
            
    if export_time_start and export_time_end:
        try:
            start_t = datetime.strptime(export_time_start, '%H:%M').time()
            end_t = datetime.strptime(export_time_end, '%H:%M').time()
            logs = logs.filter(timestamp__time__range=(start_t, end_t))
            title_suffix += f" ({export_time_start}-{export_time_end})"
        except ValueError:
            pass
            
    if export_person and export_person != 'All':
        if export_person.lower() == 'unknown':
            logs = logs.filter(status='UNKNOWN')
            title_suffix += " - Unknown Persons"
        else:
            logs = logs.filter(person_name__iexact=export_person)
            title_suffix += f" - {export_person}"
        filename = f"Surveillance_Report_{export_person}.xlsx"
        
    daily_reports = logs.annotate(date=TruncDate('timestamp')).values('date', 'person_name', 'status').annotate(
        entry_time=Min('timestamp'),
        exit_time=Max('timestamp'),
        frequency=Count('id'),
        max_confidence=Max('confidence')
    ).order_by('-date', '-entry_time')
    
    wb = Workbook()
    
    if export_mode == 'separated':
        # Group by person name
        unique_persons = set()
        for rep in daily_reports:
            p_name = rep['person_name'] if rep['person_name'] else "Unknown Person"
            unique_persons.add(p_name)
            
        if not unique_persons:
            ws = wb.active
            ws.title = "No Records"
            ws.append(["No records found matching filters."])
        else:
            for i, p_name in enumerate(sorted(list(unique_persons))):
                sheet_title = p_name[:30].replace('[','').replace(']','').replace('*','').replace('?','').replace(':','').replace('/','').replace('\\','')
                if i == 0:
                    ws = wb.active
                    ws.title = sheet_title
                else:
                    ws = wb.create_sheet(title=sheet_title)
                
                person_reports = [r for r in daily_reports if (r['person_name'] if r['person_name'] else "Unknown Person") == p_name]
                _fill_excel_worksheet(ws, person_reports, f"Smart Surveillance - {p_name} ({title_suffix})")
    else:
        # Consolidated combined sheet
        ws = wb.active
        ws.title = "Recognition Reports"
        _fill_excel_worksheet(ws, daily_reports, f"Smart Surveillance - Face Recognition {title_suffix}")
        
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def test_alert(request):
    h, w = 480, 640
    frame = np.zeros((h, w, 3), dtype=np.uint8)
    cv.rectangle(frame, (100, 100), (540, 380), (0, 0, 255), 2)
    cv.putText(frame, "TEST STRANGER", (120, 90), cv.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2, cv.LINE_AA)
    ret, buffer = cv.imencode('.jpg', frame)
    if ret:
        send_alerts(buffer.tobytes(), 'Unknown Test Target', False, 1, 0.54)
        return HttpResponse("Test alert sent to Telegram! Go look at Telegram and click OK / Cancel now.")
    return HttpResponse("Failed to encode frame")
